import json
import os
import time
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

#import functions
from check_functions import check_files, check_folders
#import paths
from paths import (folder_to_track, folders_collection,
                   image_files_folder, office_files_folder, video_files_folder)

file_name_list = []
file_type_list = []
original_location_list = []
location_change_list = []

class MyHandler(FileSystemEventHandler):
  i = 1
  def on_modified(self, event):
    global file_name_list
    global file_type_list
    global original_location_list
    global location_change_list
    file_type = ''
    #check files/folders in target folder
    for filename in os.listdir(folder_to_track):
      #retrieve file/folder name and extensions of the file / folder
      fileName, file_extension = os.path.splitext(str(filename))
      #set the new name by adding date and time file was moved on to its existing name
      new_name = str(fileName) + str(file_extension)
      print(fileName)
      print(file_extension)
      #set the previous name of the file as src to be replaced
      src = folder_to_track + '/' + str(fileName) + str(file_extension)
      folder_destination = check_files(file_extension,src)
      print(folder_destination)
      if os.path.isdir(src):
        file_type = 'folder'
      #save file
      file_exists = os.path.isfile(folder_destination + '/' + new_name)
      #as long as a file exists in target folder execute the following
      while file_exists:
        self.i += 1
        #set the new name again
        new_name = str(fileName) + str(file_extension)
        #save the file
        file_exists = os.path.isfile(folder_destination + '/' + new_name)
      #set the new destination the file has been moved to
      new_destination = folder_destination + '/' + new_name
      #rename the file thus saving it in its new location
      os.rename(src, new_destination)
      if not file_extension:
        file_extension = file_type
        print(file_extension)
      file_name_list.append(fileName)
      print(str(fileName) + ' added to names list \n')
      file_type_list.append(str(file_extension))
      print(str(file_extension) + ' added to type list \n')
      original_location_list.append(str(src))
      print(str(src) + ' added to src list')
      location_change_list.append(str(new_destination))
      print(str(new_destination) + ' added to destination list \n')


event_handler = MyHandler()
observer = Observer()
observer.schedule(event_handler, folder_to_track, recursive=True)
observer.start()


try:
  while True:
    time.sleep(10)
  # try:
  #   while True:
  #     time.sleep(10)
  # except KeyboardInterrupt:
  #   print("trying to save \n")
  #   folder = 'C:/Users/admin_pc/Desktop/python automations'
  #   excel_path = str(datetime.now().strftime("%d-%m-%Y")) + '.xlsx'
  #   book = Workbook()
  #   sheet = book.active
  #   book.save(excel_path)
  #   data = pd.DataFrame(np.column_stack([file_name_list, file_type_list, original_location_list, location_change_list]), columns=['File Name', 'File Type', 'From', 'To'])
  #   print(data)
  #   file_exists = False
  #   for filename in os.listdir(folder):
  #     fileName, file_extension = os.path.splitext(str(filename))
  #     if fileName == str(datetime.now().strftime("%d-%m-%Y")):
  #       print("file exists \n")
  #       data.to_excel(excel_path, mode='a', index=False)
  #       print("appended dataFrame to " + str(datetime.now().strftime("%d-%m-%Y")) + ' \n')
  #       file_exists = True
  #   if file_exists == False:
  #     print("file does not exist \n")
  #     data.to_excel(excel_path, index=False)
  #     print("created new excel sheet \n")
except KeyboardInterrupt:
  print("trying to save \n")
  folder = 'C:/Users/admin_pc/Desktop/python automations'
  excel_path = 'C:/Users/admin_pc/Desktop/python automations/' + str(datetime.now().strftime("%d-%m-%Y")) + '.xlsx'
  data = pd.DataFrame(np.column_stack([file_name_list, file_type_list, original_location_list, location_change_list]))
  data.columns=['File Name', 'File Type', 'From', 'To']
  print(data)
  # book = Workbook(write_only = True)
  # sheet = book.active
  try:
    book = load_workbook(excel_path)
    if book:
      sheet = book.active
      for row in data.itertuples():
        sheet.append(row)
      book.save(excel_path)
  except FileNotFoundError:
    data.to_excel(excel_path,sheet_name='sheet1')
  # file_exists = False
  # for filename in os.listdir(folder):
  #   fileName, file_extension = os.path.splitext(str(filename))
  #   if fileName == str(datetime.now().strftime("%d-%m-%Y")):
  #     print("file exists \n")
  #     data.to_excel(excel_path, mode='a', index=False)
  #     print("appended dataFrame to " + str(datetime.now().strftime("%d-%m-%Y")) + ' \n')
  #     file_exists = True
  # if file_exists == False:
  #   print("file does not exist \n")
  #   data.to_excel(excel_path, index=False)
  #   print("created new excel sheet \n")
  # append_df_to_excel(filename=excel_path, df=data, sheet_name='sheet1' )
  observer.stop()
observer.join()
